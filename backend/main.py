"""
PDF Master Platform - Backend Main Application
Complete backend with database integration
"""

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import OAuth2PasswordRequestForm
from typing import List, Optional
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
import os
import shutil
from pathlib import Path
import tempfile
import base64
import uuid
import json
from io import BytesIO

# PDF manipulation libraries
from pypdf import PdfReader, PdfWriter
import pdfplumber
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from PIL import Image

# Document conversion libraries
from docx import Document
from openpyxl import load_workbook

# Database imports
import models
import schemas
import auth
from database import get_db, engine

# ============================================================================
# APPLICATION SETUP
# ============================================================================

app = FastAPI(
    title="PDF Master Platform API",
    description="Comprehensive PDF editing and conversion platform with database",
    version="2.0.0"
)

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================================
# DIRECTORY SETUP
# ============================================================================

UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
TEMP_DIR = Path("temp")

for directory in [UPLOAD_DIR, OUTPUT_DIR, TEMP_DIR]:
    directory.mkdir(exist_ok=True)

# ============================================================================
# DATABASE - CREATE TABLES ON STARTUP
# ============================================================================

@app.on_event("startup")
async def startup_event():
    """Create database tables on startup"""
    models.Base.metadata.create_all(bind=engine)
    print("✅ Database tables ready!")

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def save_upload_file(upload_file: UploadFile, destination: Path):
    """Save an uploaded file to disk"""
    with open(destination, "wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)

def cleanup_file(file_path: Path):
    """Delete a file safely"""
    try:
        if file_path.exists():
            file_path.unlink()
    except Exception as e:
        print(f"Error cleaning up {file_path}: {e}")

# ============================================================================
# BASIC ENDPOINTS
# ============================================================================

@app.get("/")
async def root():
    """Root endpoint"""
    return {
        "message": "PDF Master Platform API with Database",
        "version": "2.0.0",
        "status": "running",
        "docs": "http://localhost:8000/docs"
    }

@app.get("/health")
async def health_check():
    """Health check"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "database": "connected"
    }

# ============================================================================
# AUTHENTICATION ENDPOINTS
# ============================================================================

@app.post("/api/auth/register", response_model=schemas.UserResponse)
async def register(user: schemas.UserCreate, db: Session = Depends(get_db)):
    """Register new user"""
    # Check if user exists
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    db_user = db.query(models.User).filter(models.User.username == user.username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Username already taken")
    
    # Create new user
    hashed_password = auth.get_password_hash(user.password)
    db_user = models.User(
        email=user.email,
        username=user.username,
        hashed_password=hashed_password
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user


@app.post("/api/auth/login", response_model=schemas.Token)
async def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    """Login and get access token"""
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/api/auth/me", response_model=schemas.UserResponse)
async def get_me(current_user: models.User = Depends(auth.get_current_active_user)):
    """Get current user info"""
    return current_user


# ============================================================================
# FILE MANAGEMENT WITH DATABASE
# ============================================================================

@app.get("/api/files/my-files")
async def get_my_files(
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all files for current user"""
    files = db.query(models.File).filter(models.File.user_id == current_user.id).all()
    
    return {
        "success": True,
        "count": len(files),
        "files": [
            {
                "id": f.id,
                "filename": f.original_filename,
                "size": f.file_size,
                "type": f.file_type,
                "uploaded_at": f.uploaded_at.isoformat()
            }
            for f in files
        ]
    }


@app.get("/api/operations/history")
async def get_operation_history(
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db),
    limit: int = 50
):
    """Get operation history for current user"""
    operations = db.query(models.Operation).filter(
        models.Operation.user_id == current_user.id
    ).order_by(models.Operation.created_at.desc()).limit(limit).all()
    
    return {
        "success": True,
        "count": len(operations),
        "operations": [
            {
                "id": op.id,
                "type": op.operation_type,
                "status": op.status,
                "created_at": op.created_at.isoformat(),
                "completed_at": op.completed_at.isoformat() if op.completed_at else None
            }
            for op in operations
        ]
    }


@app.delete("/api/files/{file_id}")
async def delete_file(
    file_id: int,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Delete a file"""
    file = db.query(models.File).filter(
        models.File.id == file_id,
        models.File.user_id == current_user.id
    ).first()
    
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    
    # Delete physical file
    file_path = Path(file.file_path)
    if file_path.exists():
        file_path.unlink()
    
    # Delete from database
    db.delete(file)
    db.commit()
    
    return {"success": True, "message": "File deleted successfully"}

# ============================================================================
# PDF TEXT EXTRACTION
# ============================================================================

@app.post("/api/pdf/extract-text")
async def extract_text_from_pdf(file: UploadFile = File(...)):
    """Extract text from PDF with proper spacing"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    temp_path = TEMP_DIR / f"temp_{file.filename}"
    
    try:
        save_upload_file(file, temp_path)
        
        text_content = []
        with pdfplumber.open(temp_path) as pdf:
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text(layout=True, x_tolerance=2, y_tolerance=3)
                
                if not page_text:
                    page_text = ""
                else:
                    lines = page_text.split('\n')
                    cleaned_lines = []
                    for line in lines:
                        cleaned_line = ' '.join(line.split())
                        if cleaned_line:
                            cleaned_lines.append(cleaned_line)
                    page_text = '\n'.join(cleaned_lines)
                
                text_content.append({
                    "page": i + 1,
                    "text": page_text
                })
        
        return {
            "success": True,
            "filename": file.filename,
            "pages": len(text_content),
            "content": text_content
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        cleanup_file(temp_path)

# ============================================================================
# PDF TABLE EXTRACTION
# ============================================================================

@app.post("/api/pdf/extract-tables")
async def extract_tables_from_pdf(file: UploadFile = File(...)):
    """Extract tables from PDF"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    temp_path = TEMP_DIR / f"temp_{file.filename}"
    
    try:
        save_upload_file(file, temp_path)
        
        all_tables = []
        with pdfplumber.open(temp_path) as pdf:
            for i, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                for j, table in enumerate(tables):
                    if table:
                        all_tables.append({
                            "page": i + 1,
                            "table_number": j + 1,
                            "data": table
                        })
        
        return {
            "success": True,
            "filename": file.filename,
            "tables_found": len(all_tables),
            "tables": all_tables
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        cleanup_file(temp_path)

# ============================================================================
# PDF MERGE
# ============================================================================

@app.post("/api/pdf/merge")
async def merge_pdfs(files: List[UploadFile] = File(...)):
    """Merge multiple PDFs"""
    if len(files) < 2:
        raise HTTPException(status_code=400, detail="At least 2 PDFs required")
    
    temp_files = []
    output_filename = f"merged_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    output_path = OUTPUT_DIR / output_filename
    
    try:
        writer = PdfWriter()
        
        for file in files:
            if not file.filename.endswith('.pdf'):
                raise HTTPException(status_code=400, detail=f"{file.filename} is not a PDF")
            
            temp_path = TEMP_DIR / f"temp_{file.filename}"
            save_upload_file(file, temp_path)
            temp_files.append(temp_path)
            
            reader = PdfReader(temp_path)
            for page in reader.pages:
                writer.add_page(page)
        
        with open(output_path, "wb") as output_file:
            writer.write(output_file)
        
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/pdf"
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        for temp_file in temp_files:
            cleanup_file(temp_file)

# ============================================================================
# PDF SPLIT
# ============================================================================

@app.post("/api/pdf/split")
async def split_pdf(file: UploadFile = File(...)):
    """Split PDF into pages"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    temp_path = TEMP_DIR / f"temp_{file.filename}"
    
    try:
        save_upload_file(file, temp_path)
        reader = PdfReader(temp_path)
        
        output_files = []
        base_name = file.filename.replace('.pdf', '')
        
        for i, page in enumerate(reader.pages):
            writer = PdfWriter()
            writer.add_page(page)
            
            output_filename = f"{base_name}_page_{i+1}.pdf"
            output_path = OUTPUT_DIR / output_filename
            
            with open(output_path, "wb") as output_file:
                writer.write(output_file)
            
            output_files.append({
                "page": i + 1,
                "filename": output_filename,
                "path": str(output_path)
            })
        
        return {
            "success": True,
            "original_filename": file.filename,
            "total_pages": len(output_files),
            "files": output_files
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        cleanup_file(temp_path)

# ============================================================================
# PDF ROTATE
# ============================================================================

@app.post("/api/pdf/rotate")
async def rotate_pdf(
    file: UploadFile = File(...),
    angle: int = Form(90),
    pages: Optional[str] = Form("all")
):
    """Rotate PDF pages"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    if angle not in [90, 180, 270, -90]:
        raise HTTPException(status_code=400, detail="Angle must be 90, 180, 270, or -90")
    
    temp_path = TEMP_DIR / f"temp_{file.filename}"
    output_filename = f"rotated_{file.filename}"
    output_path = OUTPUT_DIR / output_filename
    
    try:
        save_upload_file(file, temp_path)
        reader = PdfReader(temp_path)
        writer = PdfWriter()
        
        pages_to_rotate = []
        if pages and pages.lower() != "all":
            pages_to_rotate = [int(p.strip()) - 1 for p in pages.split(',')]
        else:
            pages_to_rotate = list(range(len(reader.pages)))
        
        for i, page in enumerate(reader.pages):
            if i in pages_to_rotate:
                page.rotate(angle)
            writer.add_page(page)
        
        with open(output_path, "wb") as output_file:
            writer.write(output_file)
        
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/pdf"
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        cleanup_file(temp_path)

# ============================================================================
# PDF WATERMARK
# ============================================================================

@app.post("/api/pdf/add-watermark")
async def add_watermark(
    file: UploadFile = File(...),
    watermark_text: str = Form(...)
):
    """Add watermark to PDF"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    temp_path = TEMP_DIR / f"temp_{file.filename}"
    watermark_path = TEMP_DIR / "watermark.pdf"
    output_filename = f"watermarked_{file.filename}"
    output_path = OUTPUT_DIR / output_filename
    
    try:
        save_upload_file(file, temp_path)
        
        c = canvas.Canvas(str(watermark_path), pagesize=letter)
        width, height = letter
        
        c.setFont("Helvetica", 50)
        c.setFillColorRGB(0.5, 0.5, 0.5, alpha=0.3)
        c.saveState()
        c.translate(width/2, height/2)
        c.rotate(45)
        c.drawCentredString(0, 0, watermark_text)
        c.restoreState()
        c.save()
        
        watermark = PdfReader(watermark_path).pages[0]
        reader = PdfReader(temp_path)
        writer = PdfWriter()
        
        for page in reader.pages:
            page.merge_page(watermark)
            writer.add_page(page)
        
        with open(output_path, "wb") as output_file:
            writer.write(output_file)
        
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/pdf"
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        cleanup_file(temp_path)
        cleanup_file(watermark_path)

# ============================================================================
# PDF ENCRYPT
# ============================================================================

@app.post("/api/pdf/encrypt")
async def encrypt_pdf(
    file: UploadFile = File(...),
    password: str = Form(...)
):
    """Encrypt PDF with password"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    temp_path = TEMP_DIR / f"temp_{file.filename}"
    output_filename = f"encrypted_{file.filename}"
    output_path = OUTPUT_DIR / output_filename
    
    try:
        save_upload_file(file, temp_path)
        
        reader = PdfReader(temp_path)
        writer = PdfWriter()
        
        for page in reader.pages:
            writer.add_page(page)
        
        writer.encrypt(password)
        
        with open(output_path, "wb") as output_file:
            writer.write(output_file)
        
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/pdf"
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        cleanup_file(temp_path)

# ============================================================================
# FILE TO PDF CONVERSION
# ============================================================================

@app.post("/api/convert/to-pdf")
async def convert_to_pdf(file: UploadFile = File(...)):
    """Convert files to PDF"""
    filename = file.filename.lower()
    output_filename = f"{os.path.splitext(file.filename)[0]}.pdf"
    output_path = OUTPUT_DIR / output_filename
    temp_path = TEMP_DIR / file.filename
    
    try:
        save_upload_file(file, temp_path)
        
        if filename.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff')):
            image = Image.open(temp_path)
            image = image.convert('RGB')
            image.save(output_path, 'PDF')
        
        elif filename.endswith('.docx'):
            doc = Document(temp_path)
            c = canvas.Canvas(str(output_path), pagesize=letter)
            width, height = letter
            y_position = height - 50
            
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    c.drawString(50, y_position, paragraph.text[:100])
                    y_position -= 20
                    if y_position < 50:
                        c.showPage()
                        y_position = height - 50
            c.save()
        
        elif filename.endswith('.txt'):
            with open(temp_path, 'r', encoding='utf-8') as txt_file:
                content = txt_file.read()
            
            c = canvas.Canvas(str(output_path), pagesize=letter)
            width, height = letter
            y_position = height - 50
            
            lines = content.split('\n')
            for line in lines:
                if line.strip():
                    c.drawString(50, y_position, line[:100])
                    y_position -= 20
                    if y_position < 50:
                        c.showPage()
                        y_position = height - 50
            c.save()
        
        elif filename.endswith('.xlsx'):
            wb = load_workbook(temp_path)
            ws = wb.active
            
            c = canvas.Canvas(str(output_path), pagesize=letter)
            width, height = letter
            y_position = height - 50
            
            for row in ws.iter_rows(values_only=True):
                row_text = ' | '.join([str(cell) if cell else '' for cell in row])
                if row_text.strip():
                    c.drawString(50, y_position, row_text[:100])
                    y_position -= 20
                    if y_position < 50:
                        c.showPage()
                        y_position = height - 50
            c.save()
        
        else:
            raise HTTPException(status_code=400, detail="Unsupported format")
        
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/pdf"
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        cleanup_file(temp_path)

# ============================================================================
# ADVANCED PDF EDITING
# ============================================================================

@app.post("/api/pdf/get-pages-with-text-locations")
async def get_pdf_pages_with_text_locations(file: UploadFile = File(...)):
    """Get PDF pages with text locations"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    unique_id = str(uuid.uuid4())[:8]
    temp_path = TEMP_DIR / f"temp_{unique_id}_{file.filename}"
    
    try:
        import fitz
        save_upload_file(file, temp_path)
        
        pdf_document = fitz.open(str(temp_path))
        
        pages_data = []
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_bytes = pix.tobytes("png")
            img_str = base64.b64encode(img_bytes).decode()
            
            text_blocks = []
            blocks = page.get_text("dict")["blocks"]
            
            for block in blocks:
                if "lines" in block:
                    for line in block["lines"]:
                        for span in line["spans"]:
                            bbox = span["bbox"]
                            text_blocks.append({
                                "text": span["text"],
                                "x": bbox[0] * 2,
                                "y": bbox[1] * 2,
                                "width": (bbox[2] - bbox[0]) * 2,
                                "height": (bbox[3] - bbox[1]) * 2,
                                "fontSize": span["size"],
                                "fontName": span["font"],
                                "color": span["color"],
                            })
            
            pages_data.append({
                "page": page_num + 1,
                "width": pix.width,
                "height": pix.height,
                "image": f"data:image/png;base64,{img_str}",
                "textBlocks": text_blocks
            })
        
        pdf_document.close()
        
        return {
            "success": True,
            "filename": file.filename,
            "pages": pages_data
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        cleanup_file(temp_path)

@app.post("/api/pdf/edit-complete")
async def edit_pdf_complete(
    file: UploadFile = File(...),
    edits: str = Form(...),
    images: List[UploadFile] = File(default=[])
):
    """Complete PDF editing with all features"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    unique_id = str(uuid.uuid4())[:8]
    temp_path = TEMP_DIR / f"temp_{unique_id}_{file.filename}"
    output_filename = f"edited_{file.filename}"
    output_path = OUTPUT_DIR / output_filename
    
    pdf_document = None
    
    try:
        import fitz
        
        save_upload_file(file, temp_path)
        edits_data = json.loads(edits)
        
        image_paths = {}
        for img in images:
            img_path = TEMP_DIR / f"{unique_id}_{img.filename}"
            save_upload_file(img, img_path)
            image_paths[img.filename] = img_path
        
        pdf_document = fitz.open(str(temp_path))
        
        for edit in edits_data:
            page_num = edit.get('page', 1) - 1
            if page_num >= len(pdf_document):
                continue
                
            page = pdf_document[page_num]
            edit_type = edit.get('type', 'text')
            
            if edit_type == 'text':
                action = edit.get('action', 'add')
                
                if action in ['replace', 'delete']:
                    remove_area = edit.get('removeArea', {})
                    if remove_area:
                        x = remove_area.get('x', 0) / 2
                        y = remove_area.get('y', 0) / 2
                        width = remove_area.get('width', 0) / 2
                        height = remove_area.get('height', 0) / 2
                        
                        rect = fitz.Rect(x, y, x + width, y + height)
                        page.add_redact_annot(rect, fill=(1, 1, 1))
                        page.apply_redactions()
                
                if action in ['replace', 'add']:
                    text = edit.get('text', '')
                    if not text:
                        continue
                        
                    x = edit.get('x', 0) / 2
                    y = edit.get('y', 0) / 2
                    font_size = edit.get('fontSize', 12)
                    bold = edit.get('bold', False)
                    italic = edit.get('italic', False)
                    
                    if bold and italic:
                        fitz_font = 'hebo'
                    elif bold:
                        fitz_font = 'hebo'
                    elif italic:
                        fitz_font = 'heit'
                    else:
                        fitz_font = 'helv'
                    
                    color = edit.get('color', '#000000')
                    try:
                        rgb = tuple(int(color.lstrip('#')[i:i+2], 16) / 255 for i in (0, 2, 4))
                    except:
                        rgb = (0, 0, 0)
                    
                    point = fitz.Point(x, y + font_size)
                    page.insert_text(point, text, fontname=fitz_font, fontsize=font_size, color=rgb)
            
            elif edit_type == 'image':
                image_name = edit.get('imageName')
                if image_name and image_name in image_paths:
                    x = edit.get('x', 0) / 2
                    y = edit.get('y', 0) / 2
                    width = edit.get('width', 100) / 2
                    height = edit.get('height', 100) / 2
                    
                    rect = fitz.Rect(x, y, x + width, y + height)
                    page.insert_image(rect, filename=str(image_paths[image_name]))
            
            elif edit_type == 'shape':
                shape = edit.get('shape', 'rectangle')
                x = edit.get('x', 0) / 2
                y = edit.get('y', 0) / 2
                width = edit.get('width', 100) / 2
                height = edit.get('height', 100) / 2
                
                color = edit.get('color', '#000000')
                rgb = tuple(int(color.lstrip('#')[i:i+2], 16) / 255 for i in (0, 2, 4))
                
                fill_color = edit.get('fillColor')
                fill_rgb = tuple(int(fill_color.lstrip('#')[i:i+2], 16) / 255 for i in (0, 2, 4)) if fill_color else None
                
                stroke_width = edit.get('strokeWidth', 1)
                
                if shape == 'rectangle':
                    rect = fitz.Rect(x, y, x + width, y + height)
                    page.draw_rect(rect, color=rgb, fill=fill_rgb, width=stroke_width)
                
                elif shape == 'circle':
                    center = fitz.Point(x + width/2, y + height/2)
                    radius = min(width, height) / 2
                    page.draw_circle(center, radius, color=rgb, fill=fill_rgb, width=stroke_width)
                
                elif shape == 'line':
                    x2 = edit.get('x2', x + width) / 2
                    y2 = edit.get('y2', y + height) / 2
                    page.draw_line(fitz.Point(x, y), fitz.Point(x2, y2), color=rgb, width=stroke_width)
            
            elif edit_type == 'highlight':
                x = edit.get('x', 0) / 2
                y = edit.get('y', 0) / 2
                width = edit.get('width', 100) / 2
                height = edit.get('height', 100) / 2
                
                color = edit.get('color', '#FFFF00')
                rgb = tuple(int(color.lstrip('#')[i:i+2], 16) / 255 for i in (0, 2, 4))
                
                rect = fitz.Rect(x, y, x + width, y + height)
                highlight = page.add_highlight_annot(rect)
                highlight.set_colors(stroke=rgb)
                highlight.update()
        
        pdf_document.save(str(output_path))
        pdf_document.close()
        pdf_document = None
        
        import time
        time.sleep(0.1)
        
        for img_path in image_paths.values():
            try:
                cleanup_file(img_path)
            except:
                pass
        
        try:
            cleanup_file(temp_path)
        except:
            pass
        
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/pdf"
        )
    
    except Exception as e:
        if pdf_document is not None:
            try:
                pdf_document.close()
            except:
                pass
        
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    finally:
        if pdf_document is not None:
            try:
                pdf_document.close()
            except:
                pass

# ============================================================================
# FILE DOWNLOAD
# ============================================================================

@app.get("/api/pdf/download/{filename}")
async def download_file(filename: str):
    """Download file"""
    file_path = OUTPUT_DIR / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(path=file_path, filename=filename, media_type="application/pdf")

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)